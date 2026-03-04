"""
Generate SUS (System Usability Scale) Evaluation Excel Workbook
for IPDC-OSS Platform Thesis — Chapter 5 Usability Evaluation
Run: python generate_sus_excel.py
Output: SUS_Evaluation_IPDC_OSS.xlsx (same folder)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "SUS_Evaluation_IPDC_OSS.xlsx")

# ── Colour palette ────────────────────────────────────────────────────────────
BLUE_DARK   = "1565C0"
BLUE_MID    = "1976D2"
BLUE_LIGHT  = "BBDEFB"
BLUE_HEADER = "1E88E5"
GREEN_DARK  = "2E7D32"
GREEN_LIGHT = "C8E6C9"
ORANGE      = "E65100"
ORANGE_LIGHT= "FFE0B2"
GREY_LIGHT  = "F5F5F5"
GREY_MID    = "E0E0E0"
WHITE       = "FFFFFF"
YELLOW      = "FFF9C4"

def thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(style="thin", color="BDBDBD")
    n = None
    return Border(
        top=s if top else n,
        bottom=s if bottom else n,
        left=s if left else n,
        right=s if right else n,
    )

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="212121"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def merge_title(ws, cell_range, text, bg=BLUE_DARK, fg=WHITE, size=13):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = Font(name="Calibri", size=size, bold=True, color=fg)
    c.fill = fill(bg)
    c.alignment = center(wrap=True)

# ── PARTICIPANTS ──────────────────────────────────────────────────────────────
PARTICIPANTS = [
    ("P1", "IPDC Management Staff",   "Headquarters",  "Laptop (Chrome)"),
    ("P2", "IPDC Management Staff",   "Bole Lemi SEZ", "Desktop (Chrome)"),
    ("P3", "IPDC Management Staff",   "Hawassa SEZ",   "Laptop (Opera)"),
    ("P4", "IPDC Management Staff",   "Adama SEZ",     "Smartphone (Chrome)"),
    ("P5", "Tenant Representative",   "Hawassa SEZ",   "Smartphone (Chrome)"),
    ("P6", "Tenant Representative",   "Adama SEZ",     "Smartphone (Opera)"),
    ("P7", "Tenant Representative",   "Bole Lemi SEZ", "Laptop (Chrome)"),
    ("P8", "Tenant Representative",   "Kombolcha SEZ", "Smartphone (Chrome)"),
    ("P9", "Tenant Representative",   "Hawassa SEZ",   "Smartphone (Opera)"),
]

SUS_QUESTIONS = [
    (1,  "I think that I would like to use this system frequently.",                          "Positive"),
    (2,  "I found the system unnecessarily complex.",                                          "Negative"),
    (3,  "I thought the system was easy to use.",                                              "Positive"),
    (4,  "I think that I would need the support of a technical person to use this system.",    "Negative"),
    (5,  "I found the various functions in this system were well integrated.",                 "Positive"),
    (6,  "I thought there was too much inconsistency in this system.",                         "Negative"),
    (7,  "I would imagine that most people would learn to use this system very quickly.",      "Positive"),
    (8,  "I found the system very cumbersome to use.",                                         "Negative"),
    (9,  "I felt very confident using the system.",                                            "Positive"),
    (10, "I needed to learn a lot of things before I could get going with this system.",       "Negative"),
]

TASK_SCENARIOS = [
    ("Task 1", "Sign in and navigate the dashboard",
     "Log in using provided credentials and locate the service request list."),
    ("Task 2", "Submit a service request",
     "Create a new maintenance request, set a priority level, add a description, and submit."),
    ("Task 3", "Check token balance",
     "Locate the token dashboard and review the current balance and transaction history."),
    ("Task 4", "View an announcement",
     "Find the announcements section and read the most recent park notice."),
]

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Instructions
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Instructions"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 70
ws1.column_dimensions["D"].width = 20

merge_title(ws1, "A1:D1",
    "IPDC-OSS Platform — System Usability Scale (SUS) Evaluation Workbook",
    bg=BLUE_DARK, size=14)
ws1.row_dimensions[1].height = 36

merge_title(ws1, "A2:D2",
    "Usability Evaluation — 10 February 2026 | 9 Participants | Chapter 5",
    bg=BLUE_MID, size=11)
ws1.row_dimensions[2].height = 24

# Purpose section
ws1.merge_cells("A4:D4")
ws1["A4"].value = "PURPOSE"
ws1["A4"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
ws1["A4"].alignment = left()

ws1.merge_cells("A5:D5")
ws1["A5"].value = (
    "This workbook is the official data collection and calculation instrument for the SUS usability "
    "evaluation of the IPDC Digital One-Stop-Service (OSS) platform prototype. "
    "It accompanies the thesis: 'Design and Development of a Digital One-Stop-Service Platform "
    "for Industrial Parks in Ethiopia'."
)
ws1["A5"].font = body_font()
ws1["A5"].alignment = left(wrap=True)
ws1.row_dimensions[5].height = 48

# Sheets guide
headers = ["Sheet", "Purpose"]
ws1.merge_cells("A7:D7")
ws1["A7"].value = "WORKBOOK STRUCTURE"
ws1["A7"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
ws1["A7"].alignment = left()

sheet_guide = [
    ("Instructions",          "This sheet — purpose, SUS formula, scoring guide, task scenarios"),
    ("Participant Profiles",  "Participant IDs, roles, industrial parks, devices used (10-Feb-2026)"),
    ("SUS Raw Responses",     "Raw 1–5 ratings for all 10 SUS items from each participant — FILL THIS IN"),
    ("SUS Score Calculation", "Auto-calculated converted scores and final SUS score per participant"),
    ("Summary & Results",     "Final scores, mean, interpretation, grade, and bar chart"),
]

row = 8
for sh, desc in sheet_guide:
    ws1.merge_cells(f"B{row}:B{row}")
    ws1[f"B{row}"].value = sh
    ws1[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=BLUE_DARK)
    ws1[f"B{row}"].fill = fill(BLUE_LIGHT)
    ws1[f"B{row}"].alignment = left(wrap=False)
    ws1[f"B{row}"].border = thin_border()

    ws1.merge_cells(f"C{row}:D{row}")
    ws1[f"C{row}"].value = desc
    ws1[f"C{row}"].font = body_font()
    ws1[f"C{row}"].alignment = left()
    ws1[f"C{row}"].border = thin_border()
    ws1.row_dimensions[row].height = 18
    row += 1

# SUS Formula
row += 1
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"].value = "SUS SCORING FORMULA"
ws1[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
ws1[f"A{row}"].alignment = left()
row += 1

formula_rows = [
    ("Odd items — Positive (Q1,Q3,Q5,Q7,Q9):", "Converted score  =  Raw response  −  1"),
    ("Even items — Negative (Q2,Q4,Q6,Q8,Q10):", "Converted score  =  5  −  Raw response"),
    ("SUS Score:", "Sum of all 10 converted scores  ×  2.5"),
    ("Scale:", "0 (worst)  →  100 (best)"),
]
for label, formula in formula_rows:
    ws1[f"B{row}"].value = label
    ws1[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color="212121")
    ws1[f"B{row}"].fill = fill(YELLOW)
    ws1[f"B{row}"].alignment = left()
    ws1[f"B{row}"].border = thin_border()
    ws1.merge_cells(f"C{row}:D{row}")
    ws1[f"C{row}"].value = formula
    ws1[f"C{row}"].font = Font(name="Calibri", size=10, color="212121", bold=False)
    ws1[f"C{row}"].fill = fill(YELLOW)
    ws1[f"C{row}"].alignment = left()
    ws1[f"C{row}"].border = thin_border()
    ws1.row_dimensions[row].height = 18
    row += 1

# Score interpretation
row += 1
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"].value = "SUS SCORE INTERPRETATION (Bangor et al. / Sauro)"
ws1[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
ws1[f"A{row}"].alignment = left()
row += 1

interp = [
    ("> 85.5",      "A",  "Excellent", GREEN_DARK,   GREEN_LIGHT),
    ("72.6 – 85.5", "B",  "Good",      "1565C0",     BLUE_LIGHT),
    ("52.0 – 72.5", "C",  "OK",        "F57F17",     "FFF9C4"),
    ("51.7 – 51.9", "D",  "Poor",      "BF360C",     ORANGE_LIGHT),
    ("< 51.6",      "F",  "Awful",     "B71C1C",     "FFCDD2"),
]
for score_range, grade, label, fg, bg_c in interp:
    for col, val in zip(["B", "C", "D"], [score_range, grade, label]):
        ws1[f"{col}{row}"].value = val
        ws1[f"{col}{row}"].font = Font(name="Calibri", size=10, bold=(col=="C"), color=fg)
        ws1[f"{col}{row}"].fill = fill(bg_c)
        ws1[f"{col}{row}"].alignment = center()
        ws1[f"{col}{row}"].border = thin_border()
    ws1.row_dimensions[row].height = 16
    row += 1

# Task scenarios
row += 1
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"].value = "TASK SCENARIOS ADMINISTERED TO PARTICIPANTS"
ws1[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
ws1[f"A{row}"].alignment = left()
row += 1

for task, title, desc in TASK_SCENARIOS:
    ws1[f"B{row}"].value = task
    ws1[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws1[f"B{row}"].fill = fill(BLUE_HEADER)
    ws1[f"B{row}"].alignment = center()
    ws1[f"B{row}"].border = thin_border()

    ws1[f"C{row}"].value = title
    ws1[f"C{row}"].font = Font(name="Calibri", size=10, bold=True, color="212121")
    ws1[f"C{row}"].fill = fill(GREY_LIGHT)
    ws1[f"C{row}"].alignment = left()
    ws1[f"C{row}"].border = thin_border()

    ws1.merge_cells(f"D{row}:D{row}")
    ws1[f"D{row}"].value = desc
    ws1[f"D{row}"].font = body_font()
    ws1[f"D{row}"].fill = fill(WHITE)
    ws1[f"D{row}"].alignment = left(wrap=True)
    ws1[f"D{row}"].border = thin_border()
    ws1.row_dimensions[row].height = 30
    row += 1

# Important note
row += 1
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"].value = (
    "⚠  IMPORTANT: Enter REAL participant responses in the 'SUS Raw Responses' sheet. "
    "All SUS scores in the thesis must reflect actual collected data. "
    "Scores in 'SUS Score Calculation' and 'Summary & Results' update automatically."
)
ws1[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
ws1[f"A{row}"].fill = fill(ORANGE_LIGHT)
ws1[f"A{row}"].alignment = left(wrap=True)
ws1[f"A{row}"].border = thin_border()
ws1.row_dimensions[row].height = 36


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Participant Profiles
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Participant Profiles")
ws2.sheet_view.showGridLines = False
for col, w in zip(["A","B","C","D","E","F"], [8, 10, 28, 24, 22, 16]):
    ws2.column_dimensions[col].width = w

merge_title(ws2, "A1:F1",
    "Usability Evaluation — Participant Profiles (10 February 2026)",
    bg=BLUE_DARK, size=13)
ws2.row_dimensions[1].height = 32

# Column headers
h_labels = ["#", "ID", "Role", "Industrial Park / SEZ", "Device Used", "Date of Evaluation"]
h_cols   = ["A", "B", "C",    "D",                      "E",           "F"]
for col, lbl in zip(h_cols, h_labels):
    c = ws2[f"{col}2"]
    c.value = lbl
    c.font = header_font(size=10)
    c.fill = fill(BLUE_HEADER)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws2.row_dimensions[2].height = 20

for i, (pid, role, park, device) in enumerate(PARTICIPANTS, start=1):
    row = i + 2
    bg = GREY_LIGHT if i % 2 == 0 else WHITE
    for col, val in zip(h_cols, [i, pid, role, park, device, "10 February 2026"]):
        c = ws2[f"{col}{row}"]
        c.value = val
        c.font = body_font(size=10)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
    ws2.row_dimensions[row].height = 18

note_row = len(PARTICIPANTS) + 4
ws2.merge_cells(f"A{note_row}:F{note_row}")
ws2[f"A{note_row}"].value = (
    "Note: Participant IDs (P1–P9) are pseudonymised to protect privacy. "
    "Signed consent forms are held by the researcher and available on request."
)
ws2[f"A{note_row}"].font = Font(name="Calibri", size=9, italic=True, color="757575")
ws2[f"A{note_row}"].alignment = left()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SUS Raw Responses
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("SUS Raw Responses")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 8   # #
ws3.column_dimensions["B"].width = 8   # ID
ws3.column_dimensions["C"].width = 26  # Role
for q in range(1, 11):
    ws3.column_dimensions[get_column_letter(3 + q)].width = 10

merge_title(ws3, "A1:M1",
    "SUS Raw Responses — Enter each participant's 1–5 rating for every question",
    bg=BLUE_DARK, size=12)
ws3.row_dimensions[1].height = 30

# Question header rows
ws3.merge_cells("A2:A3"); ws3["A2"].value = "#"
ws3.merge_cells("B2:B3"); ws3["B2"].value = "ID"
ws3.merge_cells("C2:C3"); ws3["C2"].value = "Role"
for cell_ref in ["A2", "B2", "C2"]:
    c = ws3[cell_ref]
    c.font = header_font(size=10)
    c.fill = fill(BLUE_HEADER)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws3.row_dimensions[2].height = 18
ws3.row_dimensions[3].height = 60

for q_num, q_text, q_type in SUS_QUESTIONS:
    col_letter = get_column_letter(3 + q_num)
    # Row 2: Q number
    c2 = ws3[f"{col_letter}2"]
    c2.value = f"Q{q_num}"
    c2.font = header_font(size=10)
    c2.fill = fill(BLUE_HEADER if q_type == "Positive" else ORANGE)
    c2.alignment = center()
    c2.border = thin_border()
    # Row 3: question text (abbreviated)
    c3 = ws3[f"{col_letter}3"]
    c3.value = q_text[:60] + ("…" if len(q_text) > 60 else "")
    c3.font = Font(name="Calibri", size=8, bold=False, color="212121")
    c3.fill = fill(BLUE_LIGHT if q_type == "Positive" else ORANGE_LIGHT)
    c3.alignment = center(wrap=True)
    c3.border = thin_border()

# Legend row
ws3.merge_cells("A4:C4")
ws3["A4"].value = "Scale:  1 = Strongly Disagree    2 = Disagree    3 = Neutral    4 = Agree    5 = Strongly Agree"
ws3["A4"].font = Font(name="Calibri", size=9, italic=True, color="424242")
ws3["A4"].fill = fill(YELLOW)
ws3["A4"].alignment = left()
ws3["A4"].border = thin_border()
for q_num in range(1, 11):
    col_letter = get_column_letter(3 + q_num)
    ws3[f"{col_letter}4"].fill = fill(YELLOW)
    ws3[f"{col_letter}4"].border = thin_border()
    ws3[f"{col_letter}4"].value = "Blue=Positive  Orange=Negative" if q_num == 1 else ""
    ws3[f"{col_letter}4"].font = Font(name="Calibri", size=8, italic=True)
ws3.row_dimensions[4].height = 16

# Data rows — BLANK (user fills these in)
for i, (pid, role, park, device) in enumerate(PARTICIPANTS, start=1):
    row = i + 4
    bg = GREY_LIGHT if i % 2 == 0 else WHITE
    ws3[f"A{row}"].value = i
    ws3[f"B{row}"].value = pid
    ws3[f"C{row}"].value = role
    for col in ["A", "B", "C"]:
        ws3[f"{col}{row}"].font = body_font(size=10, bold=(col == "B"))
        ws3[f"{col}{row}"].fill = fill(bg)
        ws3[f"{col}{row}"].alignment = center()
        ws3[f"{col}{row}"].border = thin_border()
    for q_num in range(1, 11):
        col_letter = get_column_letter(3 + q_num)
        c = ws3[f"{col_letter}{row}"]
        c.value = None   # ← BLANK — researcher fills this in
        c.font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
        c.fill = fill(WHITE)
        c.alignment = center()
        c.border = thin_border()
    ws3.row_dimensions[row].height = 20

# Reminder
remind_row = len(PARTICIPANTS) + 6
ws3.merge_cells(f"A{remind_row}:M{remind_row}")
ws3[f"A{remind_row}"].value = (
    "⚠  Fill in each cell with the participant's actual response (integer 1–5). "
    "Leave blank cells as blank — do not enter 0. "
    "Scores in 'SUS Score Calculation' update automatically."
)
ws3[f"A{remind_row}"].font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
ws3[f"A{remind_row}"].fill = fill(ORANGE_LIGHT)
ws3[f"A{remind_row}"].alignment = left(wrap=True)
ws3[f"A{remind_row}"].border = thin_border()
ws3.row_dimensions[remind_row].height = 30


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SUS Score Calculation
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("SUS Score Calculation")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 8
ws4.column_dimensions["C"].width = 26
for q in range(1, 11):
    ws4.column_dimensions[get_column_letter(3 + q)].width = 11
ws4.column_dimensions["N"].width = 14  # raw sum
ws4.column_dimensions["O"].width = 14  # SUS score

merge_title(ws4, "A1:O1",
    "SUS Score Calculation — Converted Scores and Final SUS Score (Auto-calculated)",
    bg=BLUE_DARK, size=12)
ws4.row_dimensions[1].height = 30

# Header row 2
ws4.merge_cells("A2:A3"); ws4["A2"].value = "#"
ws4.merge_cells("B2:B3"); ws4["B2"].value = "ID"
ws4.merge_cells("C2:C3"); ws4["C2"].value = "Role"
for q_num, _, q_type in SUS_QUESTIONS:
    col_letter = get_column_letter(3 + q_num)
    ws4.merge_cells(f"{col_letter}2:{col_letter}2")
    c2 = ws4[f"{col_letter}2"]
    c2.value = f"Q{q_num} Conv."
    c2.font = header_font(size=9)
    c2.fill = fill(BLUE_HEADER if q_type == "Positive" else ORANGE)
    c2.alignment = center(wrap=True)
    c2.border = thin_border()
    c3 = ws4[f"{col_letter}3"]
    formula_hint = "R−1" if q_type == "Positive" else "5−R"
    c3.value = formula_hint
    c3.font = Font(name="Calibri", size=9, italic=True, color="424242")
    c3.fill = fill(BLUE_LIGHT if q_type == "Positive" else ORANGE_LIGHT)
    c3.alignment = center()
    c3.border = thin_border()

ws4["N2"].value = "Raw Sum"
ws4["N3"].value = "(max 40)"
ws4["O2"].value = "SUS Score"
ws4["O3"].value = "Sum × 2.5"
for cell_ref in ["A2","B2","C2","N2","O2","A3","B3","C3","N3","O3"]:
    c = ws4[cell_ref]
    if "2" in cell_ref:
        c.font = header_font(size=10)
        c.fill = fill(BLUE_HEADER)
    else:
        c.font = Font(name="Calibri", size=9, italic=True, color="424242")
        c.fill = fill(BLUE_LIGHT)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws4.row_dimensions[2].height = 20
ws4.row_dimensions[3].height = 18

# DATA ROWS — formulas referencing Raw Responses sheet
# Raw Responses data starts at row 5 (Sheet3), columns D(Q1)..M(Q10)
# Q columns in sheet3: D=Q1, E=Q2, F=Q3, G=Q4, H=Q5, I=Q6, J=Q7, K=Q8, L=Q9, M=Q10
Q_COLS = {q: get_column_letter(3 + q) for q in range(1, 11)}  # same layout in both sheets

for i, (pid, role, park, device) in enumerate(PARTICIPANTS, start=1):
    ws4_row = i + 3   # calculation sheet row
    ws3_row = i + 4   # raw responses sheet row
    bg = GREY_LIGHT if i % 2 == 0 else WHITE

    ws4[f"A{ws4_row}"].value = i
    ws4[f"B{ws4_row}"].value = pid
    ws4[f"C{ws4_row}"].value = role
    for col in ["A", "B", "C"]:
        ws4[f"{col}{ws4_row}"].font = body_font(size=10, bold=(col == "B"))
        ws4[f"{col}{ws4_row}"].fill = fill(bg)
        ws4[f"{col}{ws4_row}"].alignment = center()
        ws4[f"{col}{ws4_row}"].border = thin_border()

    conv_cols = []
    for q_num, _, q_type in SUS_QUESTIONS:
        col_letter = get_column_letter(3 + q_num)
        raw_ref = f"'SUS Raw Responses'!{col_letter}{ws3_row}"
        if q_type == "Positive":
            formula = f"=IF({raw_ref}=\"\",\"\",{raw_ref}-1)"
        else:
            formula = f"=IF({raw_ref}=\"\",\"\",5-{raw_ref})"
        c = ws4[f"{col_letter}{ws4_row}"]
        c.value = formula
        c.font = Font(name="Calibri", size=11, bold=True,
                      color=BLUE_DARK if q_type == "Positive" else ORANGE)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
        conv_cols.append(f"{col_letter}{ws4_row}")

    # Raw sum (N)
    sum_formula = "=IF(COUNTA(" + ",".join(conv_cols) + ")=10,SUM(" + ",".join(conv_cols) + "),\"\")"
    ws4[f"N{ws4_row}"].value = sum_formula
    ws4[f"N{ws4_row}"].font = Font(name="Calibri", size=11, bold=True, color="212121")
    ws4[f"N{ws4_row}"].fill = fill(YELLOW)
    ws4[f"N{ws4_row}"].alignment = center()
    ws4[f"N{ws4_row}"].border = thin_border()

    # SUS Score (O)
    ws4[f"O{ws4_row}"].value = f'=IF(N{ws4_row}="","",N{ws4_row}*2.5)'
    ws4[f"O{ws4_row}"].font = Font(name="Calibri", size=12, bold=True, color=GREEN_DARK)
    ws4[f"O{ws4_row}"].fill = fill(GREEN_LIGHT)
    ws4[f"O{ws4_row}"].alignment = center()
    ws4[f"O{ws4_row}"].border = thin_border()
    ws4.row_dimensions[ws4_row].height = 20

# Mean row
mean_row = len(PARTICIPANTS) + 5
ws4.merge_cells(f"A{mean_row}:C{mean_row}")
ws4[f"A{mean_row}"].value = "MEAN SUS SCORE"
ws4[f"A{mean_row}"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
ws4[f"A{mean_row}"].fill = fill(GREEN_DARK)
ws4[f"A{mean_row}"].alignment = center()
ws4[f"A{mean_row}"].border = thin_border()

o_data_range = f"O4:O{len(PARTICIPANTS)+3}"
ws4[f"O{mean_row}"].value = f'=IF(COUNTA({o_data_range})=9,AVERAGE({o_data_range}),"")'
ws4[f"O{mean_row}"].number_format = "0.0"
ws4[f"O{mean_row}"].font = Font(name="Calibri", size=13, bold=True, color=WHITE)
ws4[f"O{mean_row}"].fill = fill(GREEN_DARK)
ws4[f"O{mean_row}"].alignment = center()
ws4[f"O{mean_row}"].border = thin_border()
ws4.row_dimensions[mean_row].height = 24


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Summary & Results
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary & Results")
ws5.sheet_view.showGridLines = False
for col, w in zip(["A","B","C","D","E","F","G"], [8, 10, 28, 16, 14, 18, 20]):
    ws5.column_dimensions[col].width = w

merge_title(ws5, "A1:G1",
    "SUS Evaluation — Summary of Results | IPDC-OSS Platform",
    bg=BLUE_DARK, size=13)
ws5.row_dimensions[1].height = 32

# Table header
headers5 = ["#", "ID", "Role", "Device", "SUS Score", "Grade", "Interpretation"]
for col_idx, hdr in enumerate(headers5, start=1):
    col_letter = get_column_letter(col_idx)
    c = ws5[f"{col_letter}2"]
    c.value = hdr
    c.font = header_font(size=10)
    c.fill = fill(BLUE_HEADER)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws5.row_dimensions[2].height = 20

def grade_formula(score_cell):
    return (
        f'=IF({score_cell}="","",IF({score_cell}>85.5,"A",'
        f'IF({score_cell}>=72.6,"B",'
        f'IF({score_cell}>=52,"C",'
        f'IF({score_cell}>=51.7,"D","F")))))'
    )

def interp_formula(score_cell):
    return (
        f'=IF({score_cell}="","",IF({score_cell}>85.5,"Excellent",'
        f'IF({score_cell}>=72.6,"Good",'
        f'IF({score_cell}>=52,"OK",'
        f'IF({score_cell}>=51.7,"Poor","Awful")))))'
    )

for i, (pid, role, park, device) in enumerate(PARTICIPANTS, start=1):
    row = i + 2
    bg = GREY_LIGHT if i % 2 == 0 else WHITE
    calc_row = i + 3  # row in SUS Score Calculation sheet
    score_ref = f"'SUS Score Calculation'!O{calc_row}"

    values = [i, pid, role, device]
    for col_idx, val in enumerate(values, start=1):
        col_letter = get_column_letter(col_idx)
        c = ws5[f"{col_letter}{row}"]
        c.value = val
        c.font = body_font(size=10, bold=(col_idx == 2))
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

    # SUS Score (E)
    ws5[f"E{row}"].value = f"={score_ref}"
    ws5[f"E{row}"].number_format = "0.0"
    ws5[f"E{row}"].font = Font(name="Calibri", size=11, bold=True, color=GREEN_DARK)
    ws5[f"E{row}"].fill = fill(GREEN_LIGHT)
    ws5[f"E{row}"].alignment = center()
    ws5[f"E{row}"].border = thin_border()

    # Grade (F)
    ws5[f"F{row}"].value = grade_formula(f"E{row}")
    ws5[f"F{row}"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_DARK)
    ws5[f"F{row}"].fill = fill(bg)
    ws5[f"F{row}"].alignment = center()
    ws5[f"F{row}"].border = thin_border()

    # Interpretation (G)
    ws5[f"G{row}"].value = interp_formula(f"E{row}")
    ws5[f"G{row}"].font = body_font(size=10)
    ws5[f"G{row}"].fill = fill(bg)
    ws5[f"G{row}"].alignment = center()
    ws5[f"G{row}"].border = thin_border()
    ws5.row_dimensions[row].height = 20

# Mean row
mean_row5 = len(PARTICIPANTS) + 4
ws5.merge_cells(f"A{mean_row5}:D{mean_row5}")
ws5[f"A{mean_row5}"].value = "MEAN SUS SCORE"
ws5[f"A{mean_row5}"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
ws5[f"A{mean_row5}"].fill = fill(GREEN_DARK)
ws5[f"A{mean_row5}"].alignment = center()
ws5[f"A{mean_row5}"].border = thin_border()

ws5[f"E{mean_row5}"].value = f"=AVERAGE(E3:E{len(PARTICIPANTS)+2})"
ws5[f"E{mean_row5}"].number_format = "0.0"
ws5[f"E{mean_row5}"].font = Font(name="Calibri", size=13, bold=True, color=WHITE)
ws5[f"E{mean_row5}"].fill = fill(GREEN_DARK)
ws5[f"E{mean_row5}"].alignment = center()
ws5[f"E{mean_row5}"].border = thin_border()

ws5[f"F{mean_row5}"].value = grade_formula(f"E{mean_row5}")
ws5[f"F{mean_row5}"].font = Font(name="Calibri", size=12, bold=True, color=WHITE)
ws5[f"F{mean_row5}"].fill = fill(GREEN_DARK)
ws5[f"F{mean_row5}"].alignment = center()
ws5[f"F{mean_row5}"].border = thin_border()

ws5[f"G{mean_row5}"].value = interp_formula(f"E{mean_row5}")
ws5[f"G{mean_row5}"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
ws5[f"G{mean_row5}"].fill = fill(GREEN_DARK)
ws5[f"G{mean_row5}"].alignment = center()
ws5[f"G{mean_row5}"].border = thin_border()
ws5.row_dimensions[mean_row5].height = 26

# Bar Chart
chart_row = mean_row5 + 3
chart = BarChart()
chart.type = "col"
chart.title = "SUS Scores per Participant — IPDC-OSS Platform"
chart.y_axis.title = "SUS Score"
chart.x_axis.title = "Participant"
chart.style = 10
chart.height = 12
chart.width = 20

data_ref = Reference(ws5, min_col=5, min_row=2, max_row=len(PARTICIPANTS)+2)
cats_ref = Reference(ws5, min_col=2, min_row=3, max_row=len(PARTICIPANTS)+2)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "1976D2"

# Benchmark line label
ws5[f"A{chart_row}"].value = "Reference lines: 72.5 = Good threshold | 85.5 = Excellent threshold"
ws5[f"A{chart_row}"].font = Font(name="Calibri", size=9, italic=True, color="757575")
ws5.merge_cells(f"A{chart_row}:G{chart_row}")
ws5.row_dimensions[chart_row].height = 16

ws5.add_chart(chart, f"A{chart_row+1}")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"✅  Workbook saved: {OUTPUT_FILE}")
print("    Sheets: Instructions | Participant Profiles | SUS Raw Responses |")
print("            SUS Score Calculation | Summary & Results")
print()
print("Next steps:")
print("  1. Distribute the Google Form (or paper form) to your 9 participants")
print("  2. Enter their responses in the 'SUS Raw Responses' sheet (Q1–Q10, scale 1–5)")
print("  3. All scores calculate automatically")
print("  4. Update your thesis tables to match the actual collected scores")
